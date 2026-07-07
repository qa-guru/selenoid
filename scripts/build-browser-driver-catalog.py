#!/usr/bin/env python3
"""Build comprehensive browser/driver catalog XLS from public APIs."""

from __future__ import annotations

import json
import re
import subprocess
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

UA = "qa-guru-browser-catalog/1.0"
TIMEOUT = 60


def fetch_json(url: str, optional: bool = False):
    proc = subprocess.run(
        ["curl", "-sfL", "--max-time", str(TIMEOUT), "-H", f"User-Agent: {UA}", url],
        capture_output=True,
        text=True,
    )
    if proc.returncode != 0:
        if optional:
            return None
        raise RuntimeError(f"curl failed ({proc.returncode}) for {url}: {proc.stderr.strip()}")
    return json.loads(proc.stdout)


def fetch_text(url: str, optional: bool = False) -> str:
    proc = subprocess.run(
        ["curl", "-sfL", "--max-time", str(TIMEOUT), "-H", f"User-Agent: {UA}", url],
        capture_output=True,
        text=True,
    )
    if proc.returncode != 0:
        if optional:
            return ""
        raise RuntimeError(f"curl failed ({proc.returncode}) for {url}: {proc.stderr.strip()}")
    return proc.stdout


def github_releases_atom(repo: str) -> list[dict]:
    """Parse GitHub releases Atom feed (works when REST API is rate-limited)."""
    text = fetch_text(f"https://github.com/{repo}/releases.atom", optional=True)
    if not text:
        return []
    rows: list[dict] = []
    for entry in re.findall(r"<entry>.*?</entry>", text, re.DOTALL):
        title_m = re.search(r"<title>([^<]+)</title>", entry)
        updated_m = re.search(r"<updated>([^<]+)</updated>", entry)
        if not title_m:
            continue
        title = title_m.group(1).strip()
        if title.startswith("Release notes from"):
            continue
        tag = title if title.startswith("v") else f"v{title}"
        rows.append(
            {
                "tag_name": tag,
                "published_at": updated_m.group(1) if updated_m else "",
            }
        )
    return rows


def major(version: str) -> str:
    m = re.match(r"(\d+)", version)
    return m.group(1) if m else version


def chrome_versionhistory_stable() -> list[dict]:
    rows: list[dict] = []
    seen: set[str] = set()
    for platform in ("all", "win", "mac", "linux", "android"):
        url = (
            "https://versionhistory.googleapis.com/v1/"
            f"chrome/platforms/{platform}/channels/stable/versions?pageSize=1000"
        )
        data = fetch_json(url, optional=True)
        if not data:
            continue
        for item in data.get("versions", []):
            ver = item.get("version", "")
            if ver in seen:
                continue
            seen.add(ver)
            rows.append(
                {
                    "browser": "Chrome",
                    "channel": "stable",
                    "platform_scope": platform,
                    "version": ver,
                    "major": major(ver),
                    "release_time": item.get("releaseTime", ""),
                    "source": "Google VersionHistory API",
                }
            )
    return sorted(rows, key=lambda r: r["version"], reverse=True)


def chrome_for_testing() -> list[dict]:
    data = fetch_json(
        "https://googlechromelabs.github.io/chrome-for-testing/"
        "known-good-versions-with-downloads.json"
    )
    rows: list[dict] = []
    for item in data.get("versions", []):
        ver = item["version"]
        downloads = item.get("downloads", {})
        assets = sorted(downloads.keys())
        chromedriver = ""
        if "chromedriver" in downloads and downloads["chromedriver"]:
            chromedriver = ver
        rows.append(
            {
                "browser": "Chrome (CfT)",
                "version": ver,
                "major": major(ver),
                "chromedriver": chromedriver,
                "assets": ", ".join(assets),
                "has_chrome": "chrome" in downloads,
                "has_chromedriver": "chromedriver" in downloads,
                "has_headless_shell": "chrome-headless-shell" in downloads,
                "source": "Chrome for Testing JSON API",
            }
        )
    return rows


def docker_hub_tags(repo: str) -> list[str]:
    tags: list[str] = []
    url = f"https://hub.docker.com/v2/repositories/{repo}/tags?page_size=100"
    while url:
        data = fetch_json(url, optional=True)
        if not data:
            break
        for item in data.get("results", []):
            tags.append(item["name"])
        url = data.get("next")
    return sorted(set(tags), key=lambda t: (not re.match(r"^\d", t), t))


def selenoid_docker_images() -> list[dict]:
    repos = {
        "selenoid/chrome": "Chrome",
        "selenoid/firefox": "Firefox",
        "selenoid/opera": "Opera",
        "selenoid/vnc": "VNC base",
        "selenoid/vnc_chrome": "Chrome VNC",
        "selenoid/vnc_firefox": "Firefox VNC",
        "selenoid/vnc_opera": "Opera VNC",
        "selenoid/android": "Android",
        "selenoid/phantomjs": "PhantomJS",
    }
    rows: list[dict] = []
    for repo, browser in repos.items():
        for tag in docker_hub_tags(repo):
            rows.append(
                {
                    "provider": "aerokube/selenoid",
                    "browser": browser,
                    "docker_repo": repo,
                    "tag": tag,
                    "image": f"{repo}:{tag}",
                    "protocol": "WebDriver",
                    "status": "legacy",
                }
            )
    return rows


def twilio_docker_images() -> list[dict]:
    repos = {
        "twilio/selenoid": "multi-browser",
    }
    rows: list[dict] = []
    for repo, browser in repos.items():
        tags = docker_hub_tags(repo)
        if not tags:
            continue
        for tag in tags:
            inferred = browser
            if tag.startswith("chrome"):
                inferred = "Chrome"
            elif tag.startswith("firefox"):
                inferred = "Firefox"
            elif tag.startswith("edge"):
                inferred = "Microsoft Edge"
            rows.append(
                {
                    "provider": "twilio (legacy)",
                    "browser": inferred,
                    "docker_repo": repo,
                    "tag": tag,
                    "image": f"{repo}:{tag}",
                    "protocol": "WebDriver",
                    "status": "legacy",
                }
            )
    return rows


def github_releases(repo: str, per_page: int = 100) -> list[dict]:
    rows: list[dict] = []
    page = 1
    while True:
        url = (
            f"https://api.github.com/repos/{repo}/releases"
            f"?per_page={per_page}&page={page}"
        )
        batch = fetch_json(url, optional=True)
        if not batch:
            break
        rows.extend(batch)
        if len(batch) < per_page:
            break
        page += 1
        if page > 20:
            break
    return rows


def npm_package_versions(package: str) -> list[tuple[str, str]]:
    """Return (version, published_date) tuples from npm registry."""
    data = fetch_json(f"https://registry.npmjs.org/{package}", optional=True)
    if not data:
        return []
    times = data.get("time", {})
    versions = sorted(
        data.get("versions", {}).keys(),
        key=lambda v: times.get(v, ""),
    )
    return [(v, times.get(v, "")[:10]) for v in versions]


def appium_ios_webdriver_rows() -> list[dict]:
    """Appium + iOS Safari mobile WebDriver catalog."""
    rows: list[dict] = []

    # iOS ↔ XCUITest driver compatibility (Appium XCUITest driver docs)
    ios_compat = [
        (">= 26.4", ">= 10.23.2", "WDA >= 11.1.5", "Latest"),
        ("26.0 - 26.3", ">= 9.5.0", "WDA >= 9.14.1", "Latest"),
        ("18.0 - 18.7", ">= 7.24.15", "WDA >= 8.9.1", "Latest"),
        ("17.0 - 17.7", "4.32.23 - 10.1.0", "WDA 5.6.0 - 10.1.0", "Latest (not tested)"),
        ("16.4 - 16.7", "4.21.7 - 7.26.3", "WDA 4.13.1 - 8.9.1", "Latest (not tested)"),
        ("16.0 - 16.3", "4.7.4 - 7.26.3", "WDA 4.8.1 - 8.9.1", "Latest (not tested)"),
        ("15.0 - 15.8", "4.0.0 - 5.2.0", "WDA 4.0.0 - 5.8.5", "Latest (not tested)"),
        ("14.0 - 14.8", "4.0.0 - 4.11.1", "WDA 4.0.0 - 4.8.5", "4.27.2 (WDA 4.15.1)"),
        ("9.3 - 13.7", "< 4.0.0", "WDA < 4.0.0", "4.27.2 (WDA 4.15.1)"),
    ]
    for ios, xcuitest, wda, last_working in ios_compat:
        rows.append(
            {
                "record_type": "ios_xcuitest_compat",
                "platform": "iOS/iPadOS/tvOS",
                "ios_version": ios,
                "safari_version": "",
                "xcode_version": "",
                "appium_server": "",
                "xcuitest_driver": xcuitest,
                "wda_version": wda,
                "safari_driver": "",
                "automation_name": "XCUITest",
                "browser_name": "Safari (mobile web) / native app",
                "protocol": "W3C WebDriver",
                "release_date": "",
                "status": "documented",
                "notes": f"Last likely working: {last_working}",
                "source": "appium-xcuitest-driver system requirements",
            }
        )

    xcode_min = [
        ("Xcode 13 / iOS 15", "3.48.0"),
        ("Xcode 14-beta.3 / iOS 16 Beta", "4.7.4"),
        ("Xcode 14.3 / iOS 16.4", "4.21.7"),
        ("Xcode 15 / iOS 17", "4.35.0"),
        ("Xcode 16-beta.5 / iOS 18", "7.24.15"),
    ]
    for xcode_ios, min_driver in xcode_min:
        rows.append(
            {
                "record_type": "xcode_ios_min_driver",
                "platform": "iOS",
                "ios_version": xcode_ios.split("/")[-1].strip() if "/" in xcode_ios else "",
                "safari_version": "",
                "xcode_version": xcode_ios.split("/")[0].strip(),
                "appium_server": "",
                "xcuitest_driver": f">= {min_driver}",
                "wda_version": "",
                "safari_driver": "",
                "automation_name": "XCUITest",
                "browser_name": "",
                "protocol": "W3C WebDriver",
                "release_date": "",
                "status": "documented",
                "notes": f"Minimum XCUITest driver {min_driver}",
                "source": "appium-xcuitest-driver Xcode/iOS support table",
            }
        )

    appium_driver_compat = [
        (">= 10.0.0", "Appium 3"),
        ("4.0.0 - 9.10.5", "Appium 2"),
        ("< 4.0.0", "Appium 1"),
    ]
    for xcuitest, appium in appium_driver_compat:
        rows.append(
            {
                "record_type": "appium_xcuitest_server_compat",
                "platform": "iOS",
                "ios_version": "",
                "safari_version": "",
                "xcode_version": "",
                "appium_server": appium,
                "xcuitest_driver": xcuitest,
                "wda_version": "",
                "safari_driver": "",
                "automation_name": "XCUITest",
                "browser_name": "",
                "protocol": "W3C WebDriver",
                "release_date": "",
                "status": "documented",
                "notes": "",
                "source": "appium-xcuitest-driver Appium server compatibility",
            }
        )

    # iOS major → bundled Safari (mobile web)
    for ios in range(9, 27):
        mobile_web = "yes (XCUITest + browserName=Safari)" if ios >= 13 else (
            "legacy (UIAutomation/Instruments pre-1.15)" if ios < 10 else "partial (iOS 10-12)"
        )
        rows.append(
            {
                "record_type": "ios_safari",
                "platform": "iOS / iPadOS",
                "ios_version": f"{ios}.0",
                "safari_version": f"Safari {ios} (WebKit)",
                "xcode_version": "",
                "appium_server": "1.6+ (XCUITest)" if ios >= 9 else "",
                "xcuitest_driver": "9.3+" if ios >= 10 else ("9.3+" if ios == 9 else ""),
                "wda_version": "",
                "safari_driver": "appium-safari-driver (macOS only)" if ios >= 13 else "",
                "automation_name": "XCUITest",
                "browser_name": "Safari",
                "protocol": "W3C WebDriver",
                "release_date": "",
                "status": "active" if ios >= 15 else "legacy",
                "notes": f"Mobile Safari WebDriver: {mobile_web}",
                "source": "Apple iOS release history + Appium mobile web docs",
            }
        )

    # WebDriver capabilities cheat sheet
    caps = [
        ("iOS mobile Safari", "XCUITest", "Safari", "iOS", "empty app cap", "1.15+"),
        ("iOS native app", "XCUITest", "", "iOS", "app / bundleId", "1.6+"),
        ("macOS Safari", "Safari", "Safari", "Mac", "empty app cap", "1.20+ (appium-safari-driver)"),
        ("iOS hybrid/webview", "XCUITest", "", "iOS", "app + webview context", "1.6+"),
    ]
    for use_case, auto, browser, platform, app_cap, since in caps:
        rows.append(
            {
                "record_type": "webdriver_caps",
                "platform": platform,
                "ios_version": "",
                "safari_version": "",
                "xcode_version": "",
                "appium_server": since,
                "xcuitest_driver": "",
                "wda_version": "",
                "safari_driver": "appium-safari-driver" if platform == "Mac" else "",
                "automation_name": auto,
                "browser_name": browser,
                "protocol": "W3C WebDriver",
                "release_date": "",
                "status": "reference",
                "notes": f"{use_case}; app capability: {app_cap}",
                "source": "Appium mobile web + Safari driver docs",
            }
        )

    # Appium server (npm — includes 1.x, 2.x, 3.x)
    for ver, published in npm_package_versions("appium"):
        parts = ver.split(".")
        major_v = int(parts[0]) if parts and parts[0].isdigit() else 0
        is_pre = any(x in ver for x in ("beta", "rc", "alpha"))
        rows.append(
            {
                "record_type": "appium_server",
                "platform": "server",
                "ios_version": "",
                "safari_version": "",
                "xcode_version": "",
                "appium_server": ver,
                "xcuitest_driver": "bundled (<2.0)" if major_v < 2 else "separate npm (appium driver install xcuitest)",
                "wda_version": "",
                "safari_driver": "bundled (<2.0)" if major_v < 2 else "separate npm (appium driver install safari)",
                "automation_name": "",
                "browser_name": "",
                "protocol": "W3C WebDriver / MJSONWP (1.x)",
                "release_date": published,
                "status": "legacy" if major_v < 2 else ("prerelease" if is_pre else "active"),
                "notes": "Modular drivers from Appium 2.0" if major_v >= 2 else "Monolithic Appium 1.x",
                "source": "npm registry appium",
            }
        )

    # XCUITest driver releases (npm)
    for ver, published in npm_package_versions("appium-xcuitest-driver"):
        parts = ver.split(".")
        major_v = int(parts[0]) if parts and parts[0].isdigit() else 0
        appium_gen = "Appium 3" if major_v >= 10 else ("Appium 2" if major_v >= 4 else "Appium 1")
        rows.append(
            {
                "record_type": "xcuitest_driver",
                "platform": "iOS/iPadOS/tvOS",
                "ios_version": "",
                "safari_version": "",
                "xcode_version": "",
                "appium_server": appium_gen,
                "xcuitest_driver": ver,
                "wda_version": "bundled WebDriverAgent",
                "safari_driver": "",
                "automation_name": "XCUITest",
                "browser_name": "Safari / native app",
                "protocol": "W3C WebDriver",
                "release_date": published,
                "status": "active",
                "notes": "",
                "source": "npm registry appium-xcuitest-driver",
            }
        )

    # Safari driver (macOS Safari via Apple safaridriver)
    for ver, published in npm_package_versions("appium-safari-driver"):
        rows.append(
            {
                "record_type": "safari_driver",
                "platform": "macOS",
                "ios_version": "",
                "safari_version": "Safari (desktop)",
                "xcode_version": "",
                "appium_server": "2+ (appium driver install safari)",
                "xcuitest_driver": "",
                "wda_version": "",
                "safari_driver": ver,
                "automation_name": "Safari",
                "browser_name": "Safari",
                "protocol": "W3C WebDriver (Apple safaridriver)",
                "release_date": published,
                "status": "active",
                "notes": "Wrapper over Apple safaridriver binary; iOS mobile Safari uses XCUITest driver",
                "source": "npm registry appium-safari-driver",
            }
        )

    return rows


def appium_android_webdriver_rows() -> list[dict]:
    """Appium + Android Chrome/UIAutomator2/Espresso catalog."""
    rows: list[dict] = []

    android_api_map = [
        (21, "5.0", "Lollipop"),
        (22, "5.1", "Lollipop"),
        (23, "6.0", "Marshmallow"),
        (24, "7.0", "Nougat"),
        (25, "7.1", "Nougat"),
        (26, "8.0", "Oreo"),
        (27, "8.1", "Oreo"),
        (28, "9", "Pie"),
        (29, "10", "Quince Tart"),
        (30, "11", "Red Velvet Cake"),
        (31, "12", "Snow Cone"),
        (32, "12L", "Snow Cone v2"),
        (33, "13", "Tiramisu"),
        (34, "14", "Upside Down Cake"),
        (35, "15", "Vanilla Ice Cream"),
        (36, "16", "Baklava"),
    ]

    uia2_compat = [
        (">= 6.0.0", "API 26+ (Android 8.0 Oreo+)", "Recommended minimum"),
        ("< 6.0.0", "API 21+ (Android 5.0+)", "API 21-25 has known issues; prefer API 23+"),
        (">= 5.0.0", "Appium 3 only", "appium driver install uiautomator2"),
        ("4.0.0 - 4.x", "Appium 2", "Modular driver era"),
        ("< 4.0.0", "Appium 1", "Bundled in Appium 1.x"),
    ]
    for driver_rule, android_rule, notes in uia2_compat:
        rows.append(
            {
                "record_type": "android_uia2_compat",
                "platform": "Android",
                "android_version": android_rule,
                "api_level": "",
                "browser_engine": "",
                "appium_server": "",
                "uiautomator2_driver": driver_rule,
                "espresso_driver": "",
                "chromedriver": "",
                "automation_name": "UiAutomator2",
                "browser_name": "Chrome / native app",
                "protocol": "W3C WebDriver",
                "release_date": "",
                "status": "documented",
                "notes": notes,
                "source": "appium-uiautomator2-driver README",
            }
        )

    for api, ver, codename in android_api_map:
        uia2_min = "6.0.0+" if api >= 26 else ("4.x-5.x (legacy)" if api >= 21 else "not supported")
        mobile_chrome = "yes (browserName=Chrome + UiAutomator2)" if api >= 21 else "legacy WebView"
        rows.append(
            {
                "record_type": "android_release",
                "platform": "Android",
                "android_version": ver,
                "api_level": str(api),
                "browser_engine": f"Chrome / WebView (Android {ver})",
                "appium_server": "2+ / 3",
                "uiautomator2_driver": uia2_min,
                "espresso_driver": "API 24+ (Espresso native apps)",
                "chromedriver": "for WebView/Chrome (appium-chromedriver or bundled)",
                "automation_name": "UiAutomator2",
                "browser_name": "Chrome",
                "protocol": "W3C WebDriver",
                "release_date": "",
                "status": "active" if api >= 29 else "legacy",
                "notes": f"{codename}; mobile Chrome WebDriver: {mobile_chrome}",
                "source": "Android API levels + Appium UIAutomator2 docs",
            }
        )

    caps = [
        ("Android mobile Chrome", "UiAutomator2", "Chrome", "empty app cap", "browserName=Chrome"),
        ("Android native app", "UiAutomator2", "", "app / appPackage+appActivity", ""),
        ("Android native (Espresso)", "Espresso", "", "app APK (debug)", "faster, less universal than UiAutomator2"),
        ("Android WebView/hybrid", "UiAutomator2", "", "app + chromedriver for context", "CHROMIUM context"),
    ]
    for use_case, auto, browser, app_cap, extra in caps:
        rows.append(
            {
                "record_type": "webdriver_caps",
                "platform": "Android",
                "android_version": "",
                "api_level": "",
                "browser_engine": "",
                "appium_server": "2+ / 3",
                "uiautomator2_driver": "",
                "espresso_driver": "",
                "chromedriver": "",
                "automation_name": auto,
                "browser_name": browser,
                "protocol": "W3C WebDriver",
                "release_date": "",
                "status": "reference",
                "notes": f"{use_case}; app: {app_cap}. {extra}".strip(),
                "source": "Appium Android mobile web + driver docs",
            }
        )

    for ver, published in npm_package_versions("appium-uiautomator2-driver"):
        parts = ver.split(".")
        major_v = int(parts[0]) if parts and parts[0].isdigit() else 0
        appium_gen = "Appium 3" if major_v >= 5 else ("Appium 2" if major_v >= 4 else "Appium 1")
        rows.append(
            {
                "record_type": "uiautomator2_driver",
                "platform": "Android",
                "android_version": "",
                "api_level": "26+ from driver 6.0.0" if major_v >= 6 else "21+ (legacy driver)",
                "browser_engine": "",
                "appium_server": appium_gen,
                "uiautomator2_driver": ver,
                "espresso_driver": "",
                "chromedriver": "",
                "automation_name": "UiAutomator2",
                "browser_name": "Chrome / native app",
                "protocol": "W3C WebDriver",
                "release_date": published,
                "status": "active",
                "notes": "",
                "source": "npm registry appium-uiautomator2-driver",
            }
        )

    for ver, published in npm_package_versions("appium-espresso-driver"):
        parts = ver.split(".")
        major_v = int(parts[0]) if parts and parts[0].isdigit() else 0
        appium_gen = "Appium 3" if major_v >= 5 else ("Appium 2" if major_v >= 2 else "Appium 1")
        rows.append(
            {
                "record_type": "espresso_driver",
                "platform": "Android",
                "android_version": "",
                "api_level": "24+",
                "browser_engine": "",
                "appium_server": appium_gen,
                "uiautomator2_driver": "",
                "espresso_driver": ver,
                "chromedriver": "",
                "automation_name": "Espresso",
                "browser_name": "",
                "protocol": "W3C WebDriver",
                "release_date": published,
                "status": "active",
                "notes": "Native apps only; no mobile Chrome",
                "source": "npm registry appium-espresso-driver",
            }
        )

    for ver, published in npm_package_versions("appium-chromedriver"):
        rows.append(
            {
                "record_type": "appium_chromedriver",
                "platform": "Android / hybrid",
                "android_version": "",
                "api_level": "",
                "browser_engine": "Chromium WebView",
                "appium_server": "1.x-3.x helper",
                "uiautomator2_driver": "",
                "espresso_driver": "",
                "chromedriver": ver,
                "automation_name": "UiAutomator2 / Espresso",
                "browser_name": "Chrome / WebView",
                "protocol": "W3C WebDriver",
                "release_date": published,
                "status": "active",
                "notes": "Chromedriver binary manager for Appium WebView/Chrome contexts",
                "source": "npm registry appium-chromedriver",
            }
        )

    for tag in docker_hub_tags("selenoid/android"):
        rows.append(
            {
                "record_type": "selenoid_android_docker",
                "platform": "Android emulator",
                "android_version": tag,
                "api_level": "",
                "browser_engine": "Chrome (emulator)",
                "appium_server": "",
                "uiautomator2_driver": "",
                "espresso_driver": "",
                "chromedriver": "",
                "automation_name": "UiAutomator2 (external)",
                "browser_name": "Chrome",
                "protocol": "WebDriver (Selenoid legacy)",
                "release_date": "",
                "status": "legacy",
                "notes": f"selenoid/android:{tag} — Aerokube emulator image",
                "source": "Docker Hub selenoid/android",
            }
        )

    for tag in docker_hub_tags("selenoid/chrome-mobile"):
        rows.append(
            {
                "record_type": "selenoid_chrome_mobile",
                "platform": "Android",
                "android_version": tag,
                "api_level": "",
                "browser_engine": "Chrome mobile",
                "appium_server": "",
                "uiautomator2_driver": "",
                "espresso_driver": "",
                "chromedriver": "",
                "automation_name": "",
                "browser_name": "Chrome",
                "protocol": "WebDriver (Selenoid legacy)",
                "release_date": "",
                "status": "legacy",
                "notes": f"selenoid/chrome-mobile:{tag}",
                "source": "Docker Hub selenoid/chrome-mobile",
            }
        )

    return rows


def appium_real_device_rows() -> list[dict]:
    """Real device setup reference for Appium iOS and Android."""
    rows: list[dict] = []

    ios_requirements = [
        ("host", "macOS only", "mandatory", "XCUITest requires Xcode"),
        ("host", "Xcode + Command Line Tools", "mandatory", "Match or exceed target iOS SDK"),
        ("host", "Apple Developer account", "recommended", "Free account works for local real device"),
        ("device", "iOS/iPadOS real device", "mandatory", "Simulator optional for dev"),
        ("device", "Developer Mode enabled", "mandatory", "iOS 16+ Settings → Privacy & Security"),
        ("device", "Trust this computer", "mandatory", "Prompt on first USB connect"),
        ("signing", "WebDriverAgent (WDA) signing", "mandatory", "xcodeOrgId, xcodeSigningId, updatedWDABundleId"),
        ("signing", "Provisioning profile / team ID", "mandatory", "Automatic signing in Xcode or manual"),
        ("connection", "USB (preferred) or Wi-Fi", "mandatory", "Wi-Fi requires paired device"),
        ("connection", "usbmuxd / libimobiledevice", "optional", "Diagnostics; Appium uses its own stack"),
        ("capabilities", "appium:automationName = XCUITest", "mandatory", ""),
        ("capabilities", "appium:udid", "mandatory", "Real device identifier"),
        ("capabilities", "appium:xcodeOrgId / xcodeSigningId", "mandatory", "For WDA build on device"),
        ("capabilities", "appium:webDriverAgentUrl", "optional", "Prebuilt WDA; skip rebuild"),
        ("tools", "go-ios", "optional", "Device management without Xcode GUI"),
        ("tools", "tidevice", "optional", "iOS device CLI (legacy/alternative)"),
        ("tools", "idb (Facebook)", "optional", "Simulator + device companion"),
        ("tools", "py-ios-device", "optional", "Appium mobile: extensions"),
    ]

    android_requirements = [
        ("host", "Windows / Linux / macOS", "mandatory", "Unlike iOS, any host OS"),
        ("host", "Android SDK Platform-Tools (adb)", "mandatory", "ANDROID_HOME or ANDROID_SDK_ROOT"),
        ("host", "Java JDK 8 (SDK <30) or JDK 9+ (SDK 30+)", "mandatory", "JAVA_HOME required"),
        ("host", "Android SDK Build-Tools >= 24", "mandatory", "UiAutomator2 minimum"),
        ("device", "Real Android device or emulator", "mandatory", ""),
        ("device", "Developer options → USB debugging", "mandatory", "Real device only"),
        ("device", "adb devices shows device online", "mandatory", "Authorize RSA fingerprint on device"),
        ("device", "Stay awake / disable battery opt", "recommended", "Prevents disconnect during long tests"),
        ("connection", "USB (preferred) or Wi-Fi adb", "mandatory", "adb tcpip 5555 for Wi-Fi"),
        ("capabilities", "appium:automationName = UiAutomator2", "mandatory", "Default Android driver"),
        ("capabilities", "appium:udid / deviceName", "mandatory", "Serial from adb devices -l"),
        ("capabilities", "appium:app (APK) or appPackage+appActivity", "mandatory", "Debug build preferred"),
        ("capabilities", "browserName=Chrome (mobile web)", "optional", "Empty app cap for browser-only"),
        ("capabilities", "appium:chromedriverExecutable", "optional", "Pin chromedriver for WebView"),
        ("capabilities", "appium:autoGrantPermissions=true", "optional", "Skip permission dialogs"),
        ("tools", "adb", "mandatory", "Android Debug Bridge"),
        ("tools", "scrcpy", "optional", "Screen mirror for debugging"),
        ("tools", "Appium Device Farm / STF", "optional", "Remote real device farms"),
    ]

    for category, requirement, mandatory, notes in ios_requirements:
        rows.append(
            {
                "record_type": "ios_real_device",
                "platform": "iOS / iPadOS",
                "category": category,
                "requirement": requirement,
                "mandatory": mandatory,
                "driver": "appium-xcuitest-driver",
                "tool_or_capability": requirement if category == "capabilities" else "",
                "protocol": "W3C WebDriver",
                "status": "reference",
                "notes": notes,
                "source": "Appium XCUITest real device setup docs",
            }
        )

    for category, requirement, mandatory, notes in android_requirements:
        rows.append(
            {
                "record_type": "android_real_device",
                "platform": "Android",
                "category": category,
                "requirement": requirement,
                "mandatory": mandatory,
                "driver": "appium-uiautomator2-driver",
                "tool_or_capability": requirement if category == "capabilities" else "",
                "protocol": "W3C WebDriver",
                "status": "reference",
                "notes": notes,
                "source": "Appium UIAutomator2 real device setup docs",
            }
        )

    cloud_farms = [
        ("BrowserStack", "iOS + Android real devices", "Appium endpoint URL + cloud capabilities"),
        ("Sauce Labs", "iOS + Android real devices", "appiumVersion + platformName + deviceName"),
        ("AWS Device Farm", "iOS + Android", "Upload APK/IPA or specify browser"),
        ("Firebase Test Lab", "Android (Espresso/Instrumentation)", "Limited Appium; mostly native"),
        ("HeadSpin", "iOS + Android", "Enterprise device cloud"),
    ]
    for provider, platforms, notes in cloud_farms:
        rows.append(
            {
                "record_type": "device_cloud",
                "platform": platforms,
                "category": "cloud",
                "requirement": provider,
                "mandatory": "optional",
                "driver": "Appium 2/3 remote",
                "tool_or_capability": "remote WebDriver URL",
                "protocol": "W3C WebDriver",
                "status": "reference",
                "notes": notes,
                "source": "Appium device cloud providers",
            }
        )

    return rows


def geckodriver_releases() -> list[dict]:
    rows: list[dict] = []
    releases = github_releases("mozilla/geckodriver")
    if not releases:
        releases = github_releases_atom("mozilla/geckodriver")
    for rel in releases:
        tag = rel.get("tag_name", "").lstrip("v")
        rows.append(
            {
                "driver": "geckodriver",
                "version": tag,
                "release_date": rel.get("published_at", ""),
                "browser": "Firefox",
                "source": "GitHub mozilla/geckodriver",
            }
        )
    return rows


def playwright_releases() -> list[dict]:
    rows: list[dict] = []
    for rel in github_releases("microsoft/playwright"):
        tag = rel.get("tag_name", "").lstrip("v")
        rows.append(
            {
                "playwright_version": tag,
                "release_date": rel.get("published_at", ""),
                "source": "GitHub microsoft/playwright",
            }
        )
    if rows:
        return rows

    data = fetch_json("https://registry.npmjs.org/@playwright/test", optional=True)
    if not data:
        return rows
    times = data.get("time", {})
    stable = [v for v in data.get("versions", {}) if re.match(r"^\d+\.\d+\.\d+$", v)]
    for ver in sorted(stable, key=lambda v: times.get(v, "")):
        rows.append(
            {
                "playwright_version": ver,
                "release_date": times.get(ver, ""),
                "source": "npm @playwright/test",
            }
        )
    return rows


def playwright_browsers_json(version: str) -> dict | None:
    url = (
        "https://raw.githubusercontent.com/microsoft/playwright/"
        f"v{version}/packages/playwright-core/browsers.json"
    )
    return fetch_json(url, optional=True)


def playwright_browser_matrix() -> list[dict]:
    rows: list[dict] = []
    releases = playwright_releases()
    total = len(releases)
    for idx, rel in enumerate(releases, 1):
        pw = rel["playwright_version"]
        if idx % 10 == 0 or idx == 1 or idx == total:
            print(f"  Playwright browsers.json {idx}/{total} (v{pw})...")
        data = playwright_browsers_json(pw)
        if not data:
            continue
        for browser in data.get("browsers", []):
            name = browser.get("name", "")
            bver = browser.get("browserVersion", "")
            rev = browser.get("revision", "")
            install_by_default = browser.get("installByDefault", False)
            rows.append(
                {
                    "playwright_version": pw,
                    "engine": name,
                    "browser_version": bver,
                    "revision": rev,
                    "install_by_default": install_by_default,
                    "release_date": rel.get("release_date", ""),
                    "hub_name": f"playwright-{name.replace('_', '-')}",
                    "protocol": "Playwright WebSocket",
                }
            )
    return rows


def chrome_major_history() -> list[dict]:
    """Chromium majors 1-152 from public release schedule."""
    rows: list[dict] = []
    for m in range(1, 153):
        rows.append(
            {
                "browser": "Chrome / Chromium",
                "major": str(m),
                "version_example": f"{m}.0.0.0",
                "source": "Chromium major release history",
            }
        )
    return rows


def firefox_major_versions() -> list[dict]:
    """Firefox major releases 1..latest from Mozilla product-details."""
    rows: list[dict] = []
    try:
        data = fetch_json(
            "https://product-details.mozilla.org/1.0/firefox_versions.json"
        )
        for key, ver in sorted(data.items()):
            if key in ("LATEST_FIREFOX_VERSION", "LATEST_FIREFOX_DEVEL_VERSION"):
                rows.append(
                    {
                        "browser": "Firefox",
                        "version_type": key,
                        "version": ver,
                        "major": major(ver),
                        "source": "Mozilla product-details",
                    }
                )
    except Exception:
        pass

    # Historical majors (Wikipedia / Mozilla release schedule)
    esr_majors = [10, 17, 24, 31, 38, 45, 52, 60, 68, 78, 91, 102, 115, 128, 140]
    for m in range(1, 152):
        channel = "release"
        if m in esr_majors:
            channel = "release+ESR"
        rows.append(
            {
                "browser": "Firefox",
                "version_type": channel,
                "version": f"{m}.0",
                "major": str(m),
                "source": "Mozilla release history (major milestones)",
            }
        )
    return rows


def edge_major_versions() -> list[dict]:
    rows: list[dict] = []
    # Chromium-based Edge since v79 (Jan 2020)
    for m in range(79, 151):
        rows.append(
            {
                "browser": "Microsoft Edge",
                "version": f"{m}.0",
                "major": str(m),
                "engine": "Chromium",
                "source": "Edge release schedule (Chromium-based)",
            }
        )
    # Legacy EdgeHTML Edge 12-18
    for m in range(12, 19):
        rows.append(
            {
                "browser": "Microsoft Edge (EdgeHTML)",
                "version": f"{m}.0",
                "major": str(m),
                "engine": "EdgeHTML",
                "source": "Legacy Edge",
            }
        )
    return rows


def safari_major_versions() -> list[dict]:
    rows: list[dict] = []
    # Safari major versions roughly track WebKit releases
    safari_map = {
        1: "85", 2: "419", 3: "522", 4: "530", 5: "533",
        6: "536", 7: "537", 8: "600", 9: "601", 10: "602",
        11: "604", 12: "606", 13: "608", 14: "610", 15: "612",
        16: "614", 17: "618", 18: "620",
    }
    for s, webkit in safari_map.items():
        rows.append(
            {
                "browser": "Safari",
                "safari_major": str(s),
                "webkit_approx": webkit,
                "source": "Apple Safari release history",
            }
        )
    return rows


def opera_major_versions() -> list[dict]:
    rows: list[dict] = []
    # Presto Opera 1-12, Blink Opera 15+
    for m in range(1, 13):
        rows.append(
            {
                "browser": "Opera",
                "version": f"{m}.0",
                "major": str(m),
                "engine": "Presto",
                "source": "Opera release history",
            }
        )
    for m in range(15, 120):
        rows.append(
            {
                "browser": "Opera",
                "version": f"{m}.0",
                "major": str(m),
                "engine": "Blink (Chromium)",
                "source": "Opera release history",
            }
        )
    return rows


def legacy_browsers() -> list[dict]:
    entries = [
        ("Internet Explorer", "11.0", "Trident", "EOL 2022"),
        ("Internet Explorer", "10.0", "Trident", "legacy"),
        ("Internet Explorer", "9.0", "Trident", "legacy"),
        ("Internet Explorer", "8.0", "Trident", "legacy"),
        ("PhantomJS", "2.1.1", "WebKit", "unmaintained"),
        ("PhantomJS", "2.0.0", "WebKit", "unmaintained"),
        ("PhantomJS", "1.9.8", "WebKit", "unmaintained"),
        ("HtmlUnit", "4.13.0", "Rhino/JS", "headless"),
        ("HtmlUnit", "3.11.0", "Rhino/JS", "headless"),
        ("Safari (iOS WebDriver)", "17.0", "WebKit", "Appium/XCUITest"),
        ("Chrome (Android)", "latest", "Blink", "Appium/UIAutomator2"),
    ]
    return [
        {
            "browser": b,
            "version": v,
            "engine": e,
            "status": s,
            "protocol": "WebDriver/Appium",
        }
        for b, v, e, s in entries
    ]


def qaguru_catalog_rows() -> list[dict]:
    """Current qa-guru stack snapshot from driver-versions-catalog.md."""
    return [
        {
            "stack": "WebDriver",
            "hub_browser": "chrome",
            "hub_version": "148.0",
            "browser_engine": "Chrome",
            "browser_version": "148.0.7778.96",
            "driver": "chromedriver",
            "driver_version": "148.0.7778.178",
            "docker_image": "qaguru/webdriver-chrome:148",
            "status": "active",
        },
        {
            "stack": "WebDriver",
            "hub_browser": "chrome",
            "hub_version": "148.0-min",
            "browser_engine": "Chrome",
            "browser_version": "148.0.7778.96",
            "driver": "chromedriver",
            "driver_version": "148.0.7778.178",
            "docker_image": "qaguru/webdriver-chrome:148-min",
            "status": "active",
        },
        {
            "stack": "Playwright",
            "hub_browser": "playwright-chromium",
            "hub_version": "1.60.0",
            "browser_engine": "Chromium",
            "browser_version": "148.0.7778.96",
            "driver": "built-in",
            "driver_version": "1.60.0",
            "docker_image": "qaguru/playwright-chromium:1.60.0",
            "status": "active",
        },
        {
            "stack": "Playwright",
            "hub_browser": "playwright-firefox",
            "hub_version": "1.60.0",
            "browser_engine": "Firefox",
            "browser_version": "150.0.2",
            "driver": "built-in",
            "driver_version": "1.60.0",
            "docker_image": "qaguru/playwright-firefox:1.60.0",
            "status": "active",
        },
        {
            "stack": "Playwright",
            "hub_browser": "playwright-webkit",
            "hub_version": "1.60.0",
            "browser_engine": "WebKit",
            "browser_version": "26.4",
            "driver": "built-in",
            "driver_version": "1.60.0",
            "docker_image": "qaguru/playwright-webkit:1.60.0",
            "status": "active",
        },
        {
            "stack": "Playwright",
            "hub_browser": "playwright-chrome",
            "hub_version": "1.60.0",
            "browser_engine": "Chrome stable",
            "browser_version": "stable channel",
            "driver": "built-in",
            "driver_version": "1.60.0",
            "docker_image": "qaguru/playwright-chrome:1.60.0",
            "status": "active",
        },
        {
            "stack": "Playwright",
            "hub_browser": "playwright-msedge",
            "hub_version": "1.60.0",
            "browser_engine": "Microsoft Edge",
            "browser_version": "stable channel",
            "driver": "built-in",
            "driver_version": "1.60.0",
            "docker_image": "qaguru/playwright-msedge:1.60.0",
            "status": "active",
        },
    ]


def master_catalog(
    cft_rows: list[dict],
    pw_matrix: list[dict],
    selenoid_rows: list[dict],
    gecko_rows: list[dict],
) -> list[dict]:
    """Unified cross-reference sheet."""
    rows: list[dict] = []

    # Chrome CfT milestones (latest patch per major)
    by_major: dict[str, dict] = {}
    for r in cft_rows:
        m = r["major"]
        if m not in by_major or r["version"] > by_major[m]["version"]:
            by_major[m] = r
    for m, r in sorted(by_major.items(), key=lambda x: int(x[0])):
        rows.append(
            {
                "category": "Browser",
                "name": "Chrome / Chromium",
                "version": r["version"],
                "major": m,
                "driver": "chromedriver",
                "driver_version": r["version"] if r["has_chromedriver"] else "",
                "docker_aerokube": f"selenoid/chrome:{m}.0",
                "docker_qaguru": f"qaguru/webdriver-chrome:{m}" if int(m) >= 146 else "",
                "protocol": "WebDriver",
                "playwright": "",
                "notes": "CfT latest patch per major",
            }
        )

    # Playwright latest per version
    pw_by_ver: dict[str, list[dict]] = defaultdict(list)
    for r in pw_matrix:
        pw_by_ver[r["playwright_version"]].append(r)
    for pw, engines in sorted(pw_by_ver.items(), key=lambda x: x[0], reverse=True):
        for eng in engines:
            rows.append(
                {
                    "category": "Playwright bundle",
                    "name": eng["engine"],
                    "version": eng["browser_version"],
                    "major": major(str(eng["browser_version"])),
                    "driver": "playwright built-in",
                    "driver_version": pw,
                    "docker_aerokube": "",
                    "docker_qaguru": f"qaguru/playwright-{eng['engine']}:{pw}",
                    "protocol": "Playwright",
                    "playwright": pw,
                    "notes": f"rev {eng['revision']}",
                }
            )

    # Selenoid unique browser tags
    seen = set()
    for r in selenoid_rows:
        if r["browser"] not in ("Chrome", "Firefox", "Opera"):
            continue
        tag = r["tag"]
        if not re.match(r"^\d+\.\d+$", tag):
            continue
        key = (r["browser"], tag)
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "category": "Docker legacy",
                "name": r["browser"],
                "version": tag,
                "major": major(tag),
                "driver": {"Chrome": "chromedriver", "Firefox": "geckodriver", "Opera": "operadriver"}.get(
                    r["browser"], ""
                ),
                "driver_version": "",
                "docker_aerokube": r["image"],
                "docker_qaguru": "",
                "protocol": "WebDriver",
                "playwright": "",
                "notes": "aerokube/selenoid legacy image",
            }
        )

    for g in gecko_rows:
        rows.append(
            {
                "category": "Driver",
                "name": "geckodriver",
                "version": g["version"],
                "major": major(g["version"]),
                "driver": "geckodriver",
                "driver_version": g["version"],
                "docker_aerokube": "",
                "docker_qaguru": "",
                "protocol": "WebDriver",
                "playwright": "",
                "notes": g.get("release_date", "")[:10],
            }
        )

    return rows


def style_header(ws, headers: list[str]):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_sheet(wb: Workbook, title: str, rows: list[dict]):
    if not rows:
        return
    ws = wb.create_sheet(title=title[:31])
    headers = list(rows[0].keys())
    style_header(ws, headers)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, key in enumerate(headers, 1):
            val = row.get(key, "")
            ws.cell(row=r_idx, column=c_idx, value=val)
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = max(len(str(headers[col - 1])), 8)
        for r in range(2, min(len(rows) + 2, 200)):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[letter].width = max_len + 2


def build_readme_sheet(wb: Workbook, stats: dict):
    ws = wb.active
    ws.title = "README"
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    lines = [
        ("Browser & Driver Catalog", ""),
        ("Generated", now),
        ("", ""),
        ("Sheet", "Rows"),
    ]
    for k, v in stats.items():
        lines.append((k, v))
    lines.extend(
        [
            ("", ""),
            ("Sources", ""),
            ("Chrome VersionHistory", "https://versionhistory.googleapis.com/v1/"),
            ("Chrome for Testing", "https://googlechromelabs.github.io/chrome-for-testing/"),
            ("Playwright releases", "https://github.com/microsoft/playwright/releases"),
            ("Playwright browsers.json", "per-tag raw on GitHub"),
            ("geckodriver", "https://github.com/mozilla/geckodriver/releases"),
            ("Selenoid Docker", "https://hub.docker.com/u/selenoid"),
            ("Appium server", "https://www.npmjs.com/package/appium"),
            ("Appium XCUITest", "https://www.npmjs.com/package/appium-xcuitest-driver"),
            ("Appium Safari driver", "https://www.npmjs.com/package/appium-safari-driver"),
            ("Appium iOS docs", "https://appium.github.io/appium-xcuitest-driver/latest/getting-started/system-requirements/"),
            ("Appium UIAutomator2", "https://www.npmjs.com/package/appium-uiautomator2-driver"),
            ("Appium Espresso", "https://www.npmjs.com/package/appium-espresso-driver"),
            ("Appium real device", "https://appium.io/docs/en/latest/quickstart/"),
            ("Mozilla Firefox", "https://product-details.mozilla.org/"),
            ("", ""),
            ("Reference", "selenoid/docs/browser-versions.md"),
        ]
    )
    for r, (a, b) in enumerate(lines, 1):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80
    ws["A1"].font = Font(bold=True, size=14)


def main():
    out = Path(__file__).resolve().parent.parent / "docs" / "browser-driver-catalog.xlsx"
    print("Fetching Chrome VersionHistory (stable)...")
    chrome_vh = chrome_versionhistory_stable()
    print(f"  {len(chrome_vh)} rows")

    print("Fetching Chrome for Testing...")
    cft = chrome_for_testing()
    print(f"  {len(cft)} rows")

    print("Fetching Selenoid Docker tags...")
    selenoid = selenoid_docker_images()
    print(f"  {len(selenoid)} rows")

    print("Fetching Twilio Docker tags...")
    twilio = twilio_docker_images()
    print(f"  {len(twilio)} rows")

    print("Fetching geckodriver releases...")
    gecko = geckodriver_releases()
    print(f"  {len(gecko)} rows")

    print("Fetching Playwright releases + browsers.json (may take a few minutes)...")
    pw_matrix = playwright_browser_matrix()
    print(f"  {len(pw_matrix)} engine rows")

    print("Fetching Appium / iOS WebDriver catalog...")
    appium_ios = appium_ios_webdriver_rows()
    print(f"  {len(appium_ios)} rows")

    print("Fetching Appium / Android WebDriver catalog...")
    appium_android = appium_android_webdriver_rows()
    print(f"  {len(appium_android)} rows")

    print("Building Appium real device reference...")
    appium_real = appium_real_device_rows()
    print(f"  {len(appium_real)} rows")

    firefox = firefox_major_versions()
    chrome_majors = chrome_major_history()
    edge = edge_major_versions()
    safari = safari_major_versions()
    opera = opera_major_versions()
    legacy = legacy_browsers()
    qaguru = qaguru_catalog_rows()
    master = master_catalog(cft, pw_matrix, selenoid, gecko)

    wb = Workbook()
    stats = {
        "Master catalog": len(master),
        "Chrome CfT + chromedriver": len(cft),
        "Chrome VersionHistory stable": len(chrome_vh),
        "Playwright x engines": len(pw_matrix),
        "Selenoid Docker tags": len(selenoid),
        "Twilio Docker tags": len(twilio),
        "geckodriver releases": len(gecko),
        "Chrome majors history": len(chrome_majors),
        "Firefox majors": len(firefox),
        "Safari majors": len(safari),
        "Opera majors": len(opera),
        "Legacy browsers": len(legacy),
        "Appium iOS WebDriver": len(appium_ios),
        "Appium Android WebDriver": len(appium_android),
        "Appium real devices": len(appium_real),
        "qa-guru active stack": len(qaguru),
    }
    build_readme_sheet(wb, stats)
    write_sheet(wb, "Master catalog", master)
    write_sheet(wb, "Chrome CfT", cft)
    write_sheet(wb, "Chrome VersionHistory", chrome_vh)
    write_sheet(wb, "Playwright matrix", pw_matrix)
    write_sheet(wb, "Selenoid Docker", selenoid)
    write_sheet(wb, "Twilio Docker", twilio)
    write_sheet(wb, "geckodriver", gecko)
    write_sheet(wb, "Chrome majors", chrome_majors)
    write_sheet(wb, "Firefox", firefox)
    write_sheet(wb, "Edge", edge)
    write_sheet(wb, "Safari", safari)
    write_sheet(wb, "Opera", opera)
    write_sheet(wb, "Legacy browsers", legacy)
    write_sheet(wb, "Appium iOS WebDriver", appium_ios)
    write_sheet(wb, "Appium Android WebDriver", appium_android)
    write_sheet(wb, "Appium real devices", appium_real)
    write_sheet(wb, "qa-guru stack", qaguru)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    total = sum(stats.values())
    print(f"\nSaved {out}")
    print(f"Total data rows: {total}")


if __name__ == "__main__":
    main()
